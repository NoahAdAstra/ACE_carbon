from openpyxl import load_workbook
import numpy as np
import os
import starting_values as val



#########################################################################################
A_matrix_stringer = [[172517.98,54088.62,0.00],
                     [54088.62,172517.98,0.00],
                     [0.00,0.00,59214.68]]

B_matrix_stringer = [[0,0,0],[0,0,0],[0,0,0]]

D_matrix_stringer = [[113389.70,62839.36,18839.82],
                     [62839.36, 88269.94, 18839.82],
                     [18839.82, 18839.82, 66541.71]]


A_matrix_invers = np.linalg.inv(A_matrix_stringer) 
B_matrix_invers = np.linalg.inv(B_matrix_stringer) 
D_matrix_invers = np.linalg.inv(D_matrix_stringer) 
###########################################################################################


#___setting up the data base__
if __name__ == '__main__': 
    os.chdir('../../03_FemResults')

femdir = os.getcwd()
print(os.getcwd())


maindir_stringer_stability= {} 
for i in range (1,4):
    maindir_stringer_stability.update({f'LC{i}':{}})
    for j in range (1,5):
        maindir_stringer_stability [f'LC{i}'].update({f'Stringer{j}':{}}) 

#___axial stress___
file_path = os.path.join(femdir, f'{val.filename_stringer_stability}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_stringer_stability]
row = 0
column = 5


for LoadCases in maindir_stringer_stability:
    row += 12
    help = -3
    for Stringer in maindir_stringer_stability[LoadCases]: 
        help += 3
        L = 0
        for i in range(0,3):
            L = L + ws.cell(row=row+i+help,column=column).value
        maindir_stringer_stability[LoadCases][Stringer].update ({'combined_axial_stress':L})
        
#___panel_stress___
file_path = os.path.join(femdir, f'{val.filename_stability_panels}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_stability_panels]
row = -15
column = 6

for LoadCases in maindir_stringer_stability:
    row += 30
    help = -6
    for Stringer in maindir_stringer_stability[LoadCases]: 
        help += 6
        L = 0
        for i in range(0,6):
            L = L + ws.cell(row=row+i+help,column=column).value
        maindir_stringer_stability[LoadCases][Stringer].update ({'combined_pannel_stress':L})

#__volume of colum___
pups = (val.stringer_dim_1 * val.stringer_thikness + (val.stringer_dim_2 * val.stringer_thikness ))
V_one_pbar = 200 * pups
V_one_pshell = 200*200*val.t_thikness 
V_one_column = (3*V_one_pbar)+(6*V_one_pshell)

#___comb stringer stress__
for LoadCases in maindir_stringer_stability:
    for Stringer,specific_stringer in maindir_stringer_stability[LoadCases].items(): 
        term_1 = V_one_pbar * specific_stringer['combined_axial_stress']
        term_2 = V_one_pshell * specific_stringer['combined_pannel_stress']
        sigma_tot = (term_1 + term_2)/V_one_column
        specific_stringer.update ({'sigma_tot':sigma_tot })

#___cripling___
b_top = val.stringer_dim_1/2
sigma_crip_top = (1.63/((b_top/val.stringer_thikness)**0.717)) * val.sigma_ult

b_bottom = val.stringer_dim_2
sigma_crip_bottom = (1.63/((b_bottom/val.stringer_thikness)**0.717)) * val.sigma_ult

sigma_crip_stringer = (sigma_crip_top*b_top*val.stringer_thikness*2 + sigma_crip_bottom*b_bottom*val.stringer_thikness)/(b_top*val.stringer_thikness*2 + b_bottom*val.stringer_thikness)




#___BUCKling___

E_hatt_z_top = ((A_matrix_stringer[0][0])*(((val.stringer_dim_1)**3)/12))
E_hatt_z_bottom = (val.stringer_dim_2/D_matrix_invers[0][0]) 
IE_z = E_hatt_z_top + E_hatt_z_bottom

I_z_top = (val.stringer_thikness*(val.stringer_dim_1**3))/12
I_z_bottom = (val.stringer_dim_2*(val.stringer_thikness**3))/12
I_z = I_z_bottom + I_z_top
A_stringer = val.stringer_thikness*val.stringer_dim_1 + val.stringer_dim_2*val.stringer_thikness

radius_of_gyr = ((I_z/A_stringer))**0.5
c = 1 #don't know why (simply supported?)
lambda_euler = (c * val.a_width) /radius_of_gyr

E_average = IE_z/I_z
lambda_crit = ((2*(np.pi**2)*E_average)/sigma_crip_stringer)**0.5

if lambda_crit > lambda_euler :
    sigma_buckling = sigma_crip_stringer - ((1/E_average) * ((sigma_crip_stringer/(2*np.pi))**2) * (lambda_euler**2))
else: print('error')

if sigma_crip_stringer < sigma_buckling : 
    sigma_buckel_crit = sigma_crip_stringer 
else: sigma_buckel_crit = sigma_buckling


#___RF buckel combined
for LoadCases in maindir_stringer_stability:
    for Stringer,specific_stringer in maindir_stringer_stability[LoadCases].items(): 
        RF_stringer_comb =abs(sigma_buckel_crit/(1.5 * specific_stringer['sigma_tot']))
        specific_stringer.update({'RF_stringer_comb':RF_stringer_comb})



