from openpyxl import load_workbook
import numpy as np
import os
import starting_values as val

#########################################################################################
A_matrix_stringer = val.A_matrix_stringer[val.person]

B_matrix_stringer = [[0,0,0],[0,0,0],[0,0,0]]

D_matrix_stringer = val.D_matrix_stringer[val.person]


A_matrix_invers = np.linalg.inv(A_matrix_stringer) 
#B_matrix_invers = np.linalg.inv(B_matrix_stringer) 
D_matrix_invers = np.linalg.inv(D_matrix_stringer) 
###########################################################################################

A_matrix_panel = val.A_matrix_panel[val.person]

B_matrix_panel = [[0,0,0],[0,0,0],[0,0,0]]

D_matrix_panel = val.D_matrix_panel[val.person]



E_homo_panel = A_matrix_panel[0][0]/val.t_thikness
########################################################################################

Q_11 = val.E_11/(1-(val.v_12*val.v_21))
Q_22 = val.E_22/(1-(val.v_12*val.v_21))
Q_12 = (val.v_21*val.E_11)/(1-(val.v_12*val.v_21))
Q_66 = val.G_12

Q = [[Q_11,Q_12,0],
     [Q_12,Q_22,0],
     [0 , 0 ,Q_66]]


E_homo_top = A_matrix_stringer[0][0]/val.stringer_thikness
G_homo_top = A_matrix_stringer[2][2]/val.stringer_thikness

E_homo_bottom = 1/(A_matrix_invers[0][0]*val.stringer_thikness)
G_homo_bottom = 1/(A_matrix_invers[2][2]*val.stringer_thikness) 

area_top = val.stringer_dim_1*val.stringer_thikness
area_bottom = val.stringer_dim_2*val.stringer_thikness

term_1 = (E_homo_top * area_top) +  (E_homo_bottom * area_bottom)
term_2 = (area_top) + (area_bottom)
E_homo_ges = term_1/term_2

term_1 = (G_homo_top * area_top) + (G_homo_bottom * area_bottom)
term_2 = area_top + area_bottom
G_homo_ges = term_1/term_2

#___setting up the data base__
if __name__ == '__main__': 
    os.chdir('../../03_FemResults')

femdir = os.getcwd()
print(os.getcwd())


maindir_stringer_strength= {} 
for i in range (1,4):
    maindir_stringer_strength.update({f'LC{i}':{}})
    for j in range (1,9):
        maindir_stringer_strength [f'LC{i}'].update({f'Ply{j}':{}}) 

deg_list = [45,-45,0,90,90,0,-45,45]

#___get values___
file_path = os.path.join(femdir, f'{val.filename_strength_stringer}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_strength_stringer]

column = 5
row = 0

for LoadCases in maindir_stringer_strength:
    row += 12
    i = 0
    for Ply,SpecificPly in maindir_stringer_strength[LoadCases].items():
        epsilon_stringer = ws.cell(row=row,column=column).value
        SpecificPly.update({'epsilon_stringer':epsilon_stringer})
        SpecificPly.update({'phi_d':deg_list[i]})
        i += 1


#___calculate sigma ply__
for LoadCases in maindir_stringer_strength:
    for Ply,SpecificPly in maindir_stringer_strength[LoadCases].items():
        phi = SpecificPly['phi_d']
        epsilon_ply_x = (np.cos(phi)**2) * SpecificPly['epsilon_stringer']
        epsilon_ply_y = (np.sin(phi)**2) * SpecificPly['epsilon_stringer']
        epsilon_ply_xy = -2*np.sin(phi)*np.cos(phi) * SpecificPly['epsilon_stringer']
        sigma_ply_x = (Q_11*epsilon_ply_x) + (Q_12*epsilon_ply_y)
        sigma_ply_y = (Q_12*epsilon_ply_x) + (Q_22*epsilon_ply_y)
        tau_ply = Q_66 * epsilon_ply_xy
        SpecificPly.update({'sigma_ply_x':sigma_ply_x})
        SpecificPly.update({'sigma_ply_y':sigma_ply_y})
        SpecificPly.update({'tau_ply':tau_ply})

for LoadCases in maindir_stringer_strength:
    for Ply,SpecificPly in maindir_stringer_strength[LoadCases].items():
        sigma_x = SpecificPly['sigma_ply_x']
        if  sigma_x >= 0:
            RF_FF = 1/(sigma_x/val.R_paralel_t)
        else:
            RF_FF = 1/(sigma_x/-val.R_paralel_c)
        
        SpecificPly.update({'RF_FF': RF_FF})
        SpecificPly.update({'E_homo_flange':E_homo_top})
        SpecificPly.update({'E_homo_web':E_homo_bottom})
        SpecificPly.update({'E_homo_panel':E_homo_panel})
        SpecificPly.update({'EI_combo':' '})

R_paralel_A = val.R_perpendicular_c/(2*(1+val.p_weird))
tau_C = val.R_shear*((1+(2*val.p_weird))**0.5)

for LoadCases in maindir_stringer_strength:
    for Ply in maindir_stringer_strength[LoadCases]:  
        sigma_y = maindir_stringer_strength[LoadCases][Ply]['sigma_ply_y']
        tau = maindir_stringer_strength[LoadCases][Ply]['tau_ply']
        if sigma_y >= 0:
            maindir_stringer_strength[LoadCases][Ply].update({'mode':'A'})
            RF_IFF = 1/(((((tau/val.R_shear)**2)+(((1-val.p_weird*(val.R_perpendicular_t/val.R_shear))**2)*((sigma_y/val.R_perpendicular_t)**2)))**0.5) + (val.p_weird*(sigma_y/val.R_shear)))

        if sigma_y < 0 and abs(sigma_y/tau) <= (R_paralel_A/abs(tau_C)):
            maindir_stringer_strength[LoadCases][Ply].update({'mode':'B'})
            RF_IFF = 1/((1/val.R_shear)*((((tau**2) + ((val.p_weird*sigma_y)**2))**2) + val.p_weird*sigma_y))


        if sigma_y < 0 and abs(tau/sigma_y) <= (abs(tau_C)/R_paralel_A):
            maindir_stringer_strength[LoadCases][Ply].update({'mode':'C'})
            RF_IFF = 1/((((tau/(2*(1+val.p_weird)*val.R_shear))**2)+((sigma_y/val.R_perpendicular_c)**2))*(val.R_perpendicular_c/-sigma_y))
        
        maindir_stringer_strength[LoadCases][Ply].update({'RF_IFF':RF_IFF})

for LoadCases in maindir_stringer_strength:
    for Ply in maindir_stringer_strength[LoadCases]:  
        RF_FF = maindir_stringer_strength[LoadCases][Ply]['RF_FF']
        RF_IFF = maindir_stringer_strength[LoadCases][Ply]['RF_IFF']

        if RF_FF < RF_IFF: 
            RF_comb = RF_FF
            maindir_stringer_strength[LoadCases][Ply].update({'mode':'FF'})
        else: 
            RF_comb = RF_IFF
        maindir_stringer_strength[LoadCases][Ply].update({'RF_comb':RF_comb})

print('helo')