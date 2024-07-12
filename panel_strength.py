from openpyxl import load_workbook
import numpy as np
import os
import starting_values as val

 
#___setting up the data base__
if __name__ == '__main__': 
    os.chdir('../../03_FemResults')

femdir = os.getcwd()
print(os.getcwd())


maindir_panel_strength= {} 
for i in range (1,4):
    maindir_panel_strength.update({f'LC{i}':{}})
    for j in range (1,9):
        maindir_panel_strength [f'LC{i}'].update({f'Ply{j}':{}}) 

#__get values for panel strength___
#___ply X___
file_path = os.path.join(femdir, f'{val.filename_strength_panels_x}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_strength_panels_x]


column = 6
row = -228
for LoadCases in maindir_panel_strength:
    row += 240
    for j in range(0,8): 
        id = {}
        id ['sigma_x'] = ws.cell(row=row+j*30,column=column).value
        maindir_panel_strength [LoadCases] [f'Ply{j+1}'] = id


#___ply Y___         
file_path = os.path.join(femdir, f'{val.filename_strength_panels_y}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_strength_panels_y]

column = 6
row = -228
for LoadCases in maindir_panel_strength:
    row += 240
    for j in range(0,8): 
        hey = ws.cell(row=row+j*30,column=column).value
        maindir_panel_strength [LoadCases] [f'Ply{j+1}'].update({'sigma_y': hey})

#___ply XY___         
file_path = os.path.join(femdir, f'{val.filename_strength_panels_xy}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_strength_panels_xy]

column = 6
row = -228
for LoadCases in maindir_panel_strength:
    row += 240
    for j in range(0,8): 
        hey = ws.cell(row=row+j*30,column=column).value
        maindir_panel_strength [LoadCases] [f'Ply{j+1}'].update({'tau': hey})

#___FF RF___
for LoadCases in maindir_panel_strength:
    for Ply in maindir_panel_strength[LoadCases]: 
        sigma_x = maindir_panel_strength[LoadCases][Ply]['sigma_x']
        if  sigma_x >= 0:
            RF_FF = 1/(sigma_x/val.R_paralel_t)
        else:
            RF_FF = 1/(sigma_x/-val.R_paralel_c)
        maindir_panel_strength[LoadCases][Ply].update({'RF_FF':RF_FF})
        
#___IFF RF___
R_paralel_A = val.R_perpendicular_c/(2*(1+val.p_weird))
tau_C = val.R_shear*((1+(2*val.p_weird))**0.5)

for LoadCases in maindir_panel_strength:
    for Ply in maindir_panel_strength[LoadCases]:  
        sigma_y = maindir_panel_strength[LoadCases][Ply]['sigma_y']
        tau = maindir_panel_strength[LoadCases][Ply]['tau']
        if sigma_y >= 0:
            maindir_panel_strength[LoadCases][Ply].update({'mode':'A'})
            RF_IFF = 1/(((((tau/val.R_shear)**2)+(((1-val.p_weird*(val.R_perpendicular_t/val.R_shear))**2)*((sigma_y/val.R_perpendicular_t)**2)))**0.5) + (val.p_weird*(sigma_y/val.R_shear)))

        if sigma_y < 0 and abs(sigma_y/tau) <= (R_paralel_A/abs(tau_C)):
            maindir_panel_strength[LoadCases][Ply].update({'mode':'B'})
            RF_IFF = 1/((1/val.R_shear)*((((tau**2) + ((val.p_weird*sigma_y)**2))**2) + val.p_weird*sigma_y))


        if sigma_y < 0 and abs(tau/sigma_y) <= (abs(tau_C)/R_paralel_A):
            maindir_panel_strength[LoadCases][Ply].update({'mode':'C'})
            RF_IFF = 1/((((tau/(2*(1+val.p_weird)*val.R_shear))**2)+((sigma_y/val.R_perpendicular_c)**2))*(val.R_perpendicular_c/-sigma_y))
        
        maindir_panel_strength[LoadCases][Ply].update({'RF_IFF':RF_IFF})

#___comb RF___        
for LoadCases in maindir_panel_strength:
    for Ply in maindir_panel_strength[LoadCases]:  
        RF_FF = maindir_panel_strength[LoadCases][Ply]['RF_FF']
        RF_IFF = maindir_panel_strength[LoadCases][Ply]['RF_IFF']

        if RF_FF < RF_IFF: 
            RF_comb = RF_FF
            maindir_panel_strength[LoadCases][Ply].update({'mode':'FF'})
        else: 
            RF_comb = RF_IFF
        maindir_panel_strength[LoadCases][Ply].update({'RF_comb':RF_comb})

        
print()



