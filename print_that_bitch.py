import os
from openpyxl import load_workbook
import starting_values as val

os.chdir('../../')
badir = os.getcwd()
os.chdir(f'{val.person}/Results')
femdir = os.getcwd()

import spast as pa_sta
import stringer_stability as st_sta
import panel_strength as pa_str
import stringer_strength as st_str

maindir_stringer_stability = st_sta.maindir_stringer_stability
maindir_stringer_strength = st_str.maindir_stringer_strength
maindir_panel_stability = pa_sta.maindir_panel_stability
maindir_panel_strength = pa_str.maindir_panel_strength

A_matrix = val.A_matrix_stringer[val.person]
B_matrix = st_sta.B_matrix_stringer
D_matrix = val.D_matrix_stringer[val.person]



os.chdir(f'../../{val.person}')
badir = os.getcwd()
file_path = os.path.join(badir, f'{val.result_file}.xlsx')
wb = load_workbook(file_path)
ws = wb['Tabelle1']

row=17
column = 1
for i in range(0,3):
    for j in range(0,3):
        ws.cell(row=row+i,column=column+j).value = A_matrix[i][j]

row=21
column = 1
for i in range(0,3):
    for j in range(0,3):
        ws.cell(row=row+i,column=column+j).value = B_matrix[i][j]

row=25
column = 1
for i in range(0,3):
    for j in range(0,3):
        ws.cell(row=row+i,column=column+j).value = D_matrix[i][j]


column = -4
for LoadCases in  maindir_panel_strength:
    row=32
    column += 6
    for Panels,SpecificPanel in maindir_panel_strength[LoadCases].items():
        ws.cell(row=row ,column=column).value = SpecificPanel['RF_FF']
        ws.cell(row=row ,column=column+1).value = SpecificPanel['RF_IFF']
        ws.cell(row=row ,column=column+2).value = SpecificPanel['mode']
        ws.cell(row=row ,column=column+3).value = SpecificPanel['RF_comb']
        row += 1


column = -4
for LoadCases in  maindir_stringer_strength:
    row=40
    column += 6
    for Panels,SpecificStringer in maindir_stringer_strength[LoadCases].items():
        ws.cell(row=row ,column=column).value = SpecificStringer['RF_FF']
        ws.cell(row=row ,column=column+1).value = SpecificStringer['RF_IFF']
        ws.cell(row=row ,column=column+2).value = SpecificStringer['mode']
        ws.cell(row=row ,column=column+3).value = SpecificStringer['RF_comb']
        row += 1


column = -6
for LoadCases in  maindir_panel_stability:
    column += 8
    row=53
    for Panels,SpecificPanel in maindir_panel_stability[LoadCases].items():
        ws.cell(row=row ,column=column).value = SpecificPanel['average_xx']
        ws.cell(row=row ,column=column+1).value = SpecificPanel['average_yy']
        ws.cell(row=row ,column=column+2).value = SpecificPanel['average_xy']
        ws.cell(row=row ,column=column+3).value = SpecificPanel['tau_crit_biax']
        ws.cell(row=row ,column=column+4).value = SpecificPanel['sigma_crit_biax']
        ws.cell(row=row ,column=column+5).value = SpecificPanel['RF_panel_buckel']
        row += 1


column = -3
for LoadCases in  maindir_stringer_stability:
    column += 5
    row=62
    for Panels,SpecificStringer in maindir_stringer_stability[LoadCases].items():
        ws.cell(row=row ,column=column).value = SpecificStringer['sigma_tot']
        ws.cell(row=row ,column=column+1).value = st_sta.sigma_crip_stringer
        ws.cell(row=row ,column=column+2).value = SpecificStringer['RF_stringer_comb']

        row += 1

column = 2

row = 69 # nice
for i in range (0,4):
    SpecificPly = maindir_stringer_strength['LC1']['Ply1']
    ws.cell(row=row+i ,column=column).value = SpecificPly['E_homo_flange']
    ws.cell(row=row+i ,column=column+1).value = SpecificPly['E_homo_web']
    ws.cell(row=row+i ,column=column+2).value = SpecificPly['E_homo_panel']
    ws.cell(row=row+i ,column=column+3).value = SpecificPly['E_homo_panel'] 
    
        

column = 7


for LoadCases in  maindir_stringer_stability: 
    row = 69 # nice
    for Panels,SpecificPly in maindir_stringer_stability[LoadCases].items():
        ws.cell(row=row ,column=column-1).value = SpecificPly['z_EC']
        ws.cell(row=row ,column=column).value = SpecificPly['IE_comb']
        ws.cell(row=row ,column=column+1).value = SpecificPly['radius']
        ws.cell(row=row ,column=column+2).value = SpecificPly['lambda']
        ws.cell(row=row ,column=column+3).value = SpecificPly['lambda_crit']
        row += 1

wb.save(f'{val.result_file}.xlsx')
