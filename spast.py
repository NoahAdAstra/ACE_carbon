from openpyxl import load_workbook
import numpy as np
import os
import starting_values as val






#___calculating the A,B,D matrixes___

Q_11 = val.E_11/(1-(val.v_12*val.v_21))
Q_22 = val.E_22/(1-(val.v_12*val.v_21))
Q_12 = (val.v_21*val.E_11)/(1-(val.v_12*val.v_21))
Q_66 = val.G_12

Q = [[Q_11,Q_12,0],
     [Q_12,Q_22,0],
     [0 , 0 ,Q_66]]

'''def T(phi_d):
    phi = np.deg2rad(phi_d)
    T = [[np.cos(phi)**2 , np.sin(phi)**2 , 2*np.sin(phi)*np.cos(phi)],
        [np.sin(phi)**2 , np.cos(phi)**2 , -2*np.sin(phi)*np.cos(phi)],
        [-1*np.sin(phi)*np.cos(phi) , np.sin(phi)*np.cos(phi), ((np.cos(phi)**2)-(np.sin(phi)**2))]]
    T_trans = np.transpose(T)
    return T_trans

def T_1(phi_d):
    phi = np.deg2rad(phi_d)
    T = [[np.cos(phi)**2 , np.sin(phi)**2 , -2*np.sin(phi)*np.cos(phi)],
        [np.sin(phi)**2 , np.cos(phi)**2 , 2*np.sin(phi)*np.cos(phi)],
        [np.sin(phi)*np.cos(phi) , -1*np.sin(phi)*np.cos(phi), ((np.cos(phi)**2)-(np.sin(phi)**2))]]
    return T

def Q_strich(phi):
    zwichen = np.dot(T_1(phi),Q)
    result = np.dot(zwichen,T(phi))
    return result

Q_ply_1 = Q_strich(45)
'''
#ply height 1.104
Z_hoehe = [-4.416,-3.312,-2.208,-1.104,0,1.104,2.208,3.312,4.416]

A_matrix_panel = [[517553.93, 162265.86, 0.00],
                  [162265.86, 517553.93, 0.00],
                  [0.00,	0.00,	177644.03]]

B_matrix_panel = [[0,0,0],[0,0,0],[0,0,0]]

D_matrix_panel = [[3061521.95, 1696662.73, 508675.08],
                  [1696662.73, 2383288.51, 508675.08],
                  [508675.08, 508675.08, 1796626.26]]




#___setting up the databank___
if __name__ == '__main__': 
    os.chdir('../../03_FemResults')
    
femdir = os.getcwd()


maindir_panel_stability= {} 
for i in range (1,4):
    maindir_panel_stability.update({f'LC{i}':{}})
    for j in range (1,6):
        maindir_panel_stability [f'LC{i}'].update({f'Panel{j}':{}}) 

#___get values for stabiliy___
file_path = os.path.join(femdir, f'{val.filename_stability_panels}.xlsx')
wb = load_workbook(file_path)
ws = wb[val.filename_stability_panels]


row = 6
column = 6

for LoadCases in maindir_panel_stability:
    for i in range (1,6):
        row += 6
        for j in range(0,6):
            id = {}
            id ['xx'] = ws.cell(row=row+j,column=column+0).value
            id ['xy'] = ws.cell(row=row+j,column=column+1).value
            id ['yy'] = ws.cell(row=row+j,column=column+2).value
            maindir_panel_stability [LoadCases] [f'Panel{i}'] [f'Element{j+1}'] = id


#___averaged values panels___
for LoadCases in maindir_panel_stability:
    for Panels in maindir_panel_stability[LoadCases]:
        xx = 0
        xy = 0
        yy = 0
        for Elements,values in maindir_panel_stability[LoadCases] [Panels].items():
            M = values ['xx']
            K = values ['xy']
            L = values ['yy']
            xx = xx + M
            xy = xy + K
            yy = yy + L
        maindir_panel_stability[LoadCases] [Panels] ['average_xx'] = xx/6
        maindir_panel_stability[LoadCases] [Panels] ['average_xy'] = xy/6
        maindir_panel_stability[LoadCases] [Panels] ['average_yy'] = yy/6


#___RF biax___
for LoadCases in maindir_panel_stability:
    for Panels in maindir_panel_stability[LoadCases]:
        sigma_x = maindir_panel_stability[LoadCases] [Panels] ['average_xx']
        sigma_y = maindir_panel_stability[LoadCases] [Panels] ['average_yy']
        tau_xy = maindir_panel_stability[LoadCases] [Panels] ['average_xy']

        beta = sigma_y/sigma_x
        alpha = val.a_width/val.b_length

        sigma_crit_biax = (np.pi**2/((val.b_length**2)*val.t_thikness))*(1/(((val.m_wave/alpha)**2)+(beta*(val.n_wave**2))))*(D_matrix_panel[0][0] * ((val.m_wave/alpha)**4) + 2*(D_matrix_panel[0][1]+D_matrix_panel[2][2])*(((val.m_wave*val.n_wave)/alpha)**2) + (D_matrix_panel[1][1] * (val.n_wave**4)))
        maindir_panel_stability[LoadCases] [Panels] ['sigma_crit_biax'] = sigma_crit_biax
        RF_biax = abs(sigma_crit_biax/(1.5*sigma_x))

        epsilon = ((D_matrix_panel[0][0]*D_matrix_panel[1][1])**0.5)/(D_matrix_panel[0][1]+2*D_matrix_panel[2][2])
        if epsilon >= 1:
            tau_crit_biax = (4/(val.t_thikness*(val.b_length**2)))*(((D_matrix_panel[0][0]*(D_matrix_panel[1][1]**3))**0.25)*(8.12+(5.05/epsilon)))
        if epsilon < 1:
            tau_crit_biax = (4/(val.t_thikness*(val.b_length**2)))*(((D_matrix_panel[1][1]*(D_matrix_panel[0][1]+2*D_matrix_panel[2][2]))**0.5)*(11.7+(0.532*epsilon)+(0.938*(epsilon**2))))
        maindir_panel_stability[LoadCases] [Panels] ['tau_crit_biax'] = tau_crit_biax
        RF_shear = abs(tau_crit_biax/ (1.5*tau_xy))
        RF_comb = 1/((1/RF_biax)+(1/RF_shear)**2)
        maindir_panel_stability[LoadCases] [Panels] ['RF_panel_buckel'] = RF_comb

#___homogonize___

'''E_homo = A_matrix_stringer[0][0]/2.944
G_homo = A_matrix_stringer[2][2]/2.944'''


print('hey')